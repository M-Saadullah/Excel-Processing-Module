import json
import openpyxl
import os
from pathlib import Path
import openpyxl.utils

def load_json_data(json_file_path):
    """Load and parse the JSON file containing cell data."""
    try:
        with open(json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        return data
    except FileNotFoundError:
        print(f"Error: JSON file '{json_file_path}' not found.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format in '{json_file_path}': {e}")
        return None

def get_merged_cell_mapping(worksheet):
    """
    Create a mapping of all cells to their top-left corner cell for merged ranges.
    
    Args:
        worksheet: openpyxl worksheet object
        
    Returns:
        dict: Mapping of cell coordinates to top-left corner coordinates
        Example: {(2, 2): (1, 1), (2, 3): (1, 1)} for merged range A1:C2
    """
    merged_mapping = {}
    
    for merged_range in worksheet.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        
        # Map all cells in the merged range to the top-left corner
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_mapping[(row, col)] = (min_row, min_col)
    
    return merged_mapping

def resolve_merged_cell_reference(cell_ref, merged_mapping):
    """
    Resolve a cell reference to the top-left corner if it's part of a merged range.
    
    Args:
        cell_ref (str): Excel cell reference (e.g., 'B47')
        merged_mapping (dict): Mapping from get_merged_cell_mapping()
        
    Returns:
        str: Resolved cell reference (top-left corner if merged, original if not)
    """
    try:
        # Convert cell reference to row, col coordinates
        cell_coord = openpyxl.utils.coordinate_to_tuple(cell_ref)
        row, col = cell_coord
        
        # Check if this cell is part of a merged range
        if (row, col) in merged_mapping:
            # Get the top-left corner coordinates
            top_left_row, top_left_col = merged_mapping[(row, col)]
            
            # Convert back to Excel cell reference
            resolved_ref = openpyxl.utils.get_column_letter(top_left_col) + str(top_left_row)
            
            if resolved_ref != cell_ref:
                print(f"  → Mapped merged cell {cell_ref} to top-left corner {resolved_ref}")
            
            return resolved_ref
        else:
            # Not part of a merged range, return original
            return cell_ref
            
    except Exception as e:
        print(f"Warning: Could not resolve cell reference '{cell_ref}': {e}")
        return cell_ref

def update_excel_sheet(workbook, sheet_name, cell_data):
    """Update specific sheet in Excel workbook with data from JSON."""
    try:
        # Get the specific sheet
        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")
            return 0
        
        worksheet = workbook[sheet_name]
        
        # Get merged cell mapping for this worksheet
        merged_mapping = get_merged_cell_mapping(worksheet)
        
        # Track updates for reporting
        updates_made = 0
        
        # Process each cell update
        for i, cell_info in enumerate(cell_data):
            try:
                # Extract cell information
                cell_ref = cell_info.get('cell_reference')
                value = cell_info.get('value')
                context = cell_info.get('context', 'No context provided')
                
                if not cell_ref or value is None or value == '':
                    print(f"Warning: Skipping incomplete cell data at index {i}: {cell_info}")
                    continue
                
                # Validate cell reference format
                if not isinstance(cell_ref, str) or len(cell_ref) < 2:
                    print(f"Warning: Invalid cell reference '{cell_ref}' at index {i}")
                    continue
                
                # Resolve merged cell reference to top-left corner if needed
                resolved_cell_ref = resolve_merged_cell_reference(cell_ref, merged_mapping)
                
                # Convert value to appropriate type
                if isinstance(value, str):
                    # Remove commas from numbers before parsing
                    clean_value = value.replace(',', '')
                    if clean_value.isdigit():
                        value = int(clean_value)
                    else:
                        try:
                            value = float(clean_value)
                        except ValueError:
                            pass  # Keep as string if not a number
                
                # Update the cell using resolved reference
                worksheet[resolved_cell_ref] = value
                updates_made += 1
                
                if resolved_cell_ref != cell_ref:
                    print(f"Updated merged cell {cell_ref} (→ {resolved_cell_ref}) in sheet '{sheet_name}' with value: {value}")
                else:
                    print(f"Updated cell {cell_ref} in sheet '{sheet_name}' with value: {value}")
                print(f"  Context: {context[:100]}{'...' if len(context) > 100 else ''}")
                
            except Exception as e:
                print(f"Error updating cell {cell_info.get('cell_reference', 'unknown')} in sheet '{sheet_name}': {e}")
                continue
        
        return updates_made
        
    except Exception as e:
        print(f"Error processing sheet '{sheet_name}': {e}")
        return 0

def clean_sheet_name_for_json(sheet_name):
    """
    Clean sheet name to match JSON file naming pattern.
    
    Examples:
    - 'revenue forecast' -> 'revenue_forecast'
    - '1. Fic. amounts (patenting) ' -> '1_Fic_amounts_patenting'
    - '2. Fic. amounts (commercialisat' -> '2_Fic_amounts_commercialisat'
    - 'Staff' -> 'Staff'
    - 'Summary' -> 'Summary'
    """
    # Remove leading/trailing whitespace
    cleaned = sheet_name.strip()
    
    # Replace spaces with underscores
    cleaned = cleaned.replace(' ', '_')
    
    # Remove special characters that might cause issues
    cleaned = cleaned.replace('.', '_')
    cleaned = cleaned.replace('(', '_')
    cleaned = cleaned.replace(')', '')
    cleaned = cleaned.replace(',', '_')
    cleaned = cleaned.replace('-', '_')
    
    # Remove multiple consecutive underscores
    while '__' in cleaned:
        cleaned = cleaned.replace('__', '_')
    
    # Remove leading/trailing underscores
    cleaned = cleaned.strip('_')
    
    return cleaned

def update_excel_from_json(input_folder_name, output_folder_name):
    """
    Update Excel workbooks from corresponding JSON files.
    
    Args:
        input_folder_name (str): Path to folder containing Excel workbooks
        output_folder_name (str): Path to folder containing JSON files
    """
    input_path = Path(input_folder_name)
    output_path = Path(output_folder_name)
    updated_folder = Path("Updated_excel_workbooks")
    
    # Create updated folder if it doesn't exist
    updated_folder.mkdir(exist_ok=True)
    
    if not input_path.exists():
        print(f"Error: Input folder '{input_folder_name}' does not exist.")
        return
    
    if not output_path.exists():
        print(f"Error: Output folder '{output_folder_name}' does not exist.")
        return
    
    # Get all Excel files from input folder
    excel_files = list(input_path.glob("*.xlsx"))
    
    if not excel_files:
        print(f"No Excel files found in '{input_folder_name}'")
        return
    
    print(f"Found {len(excel_files)} Excel files to process...")
    
    # Process each Excel file
    for excel_file in excel_files:
        print(f"\nProcessing: {excel_file.name}")
        
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(excel_file)
            workbook_updated = False
            total_updates = 0
            
            # Get all sheet names in the workbook
            sheet_names = workbook.sheetnames
            print(f"Available sheets: {sheet_names}")
            
            # Look for matching JSON files
            for sheet_name in sheet_names:
                # Create expected JSON filename: workbook_name_sheet_name.json
                # Clean sheet name to match JSON naming pattern
                sheet_name_clean = clean_sheet_name_for_json(sheet_name)
                workbook_name = excel_file.stem  # filename without extension
                json_filename = f"{workbook_name}_{sheet_name_clean}.json"
                json_file_path = output_path / json_filename
                
                if json_file_path.exists():
                    print(f"Found matching JSON file: {json_filename}")
                    
                    # Load JSON data
                    cell_data = load_json_data(json_file_path)
                    
                    if cell_data is None:
                        print(f"Warning: Could not load JSON data from '{json_filename}'")
                        continue
                    
                    # Handle both single object and array of objects
                    if isinstance(cell_data, dict):
                        cell_data = [cell_data]
                    elif not isinstance(cell_data, list):
                        print(f"Warning: JSON data in '{json_filename}' must be an object or array of objects.")
                        continue
                    
                    # Update the specific sheet
                    updates_made = update_excel_sheet(workbook, sheet_name, cell_data)
                    total_updates += updates_made
                    workbook_updated = True
                    
                    print(f"Updated {updates_made} cells in sheet '{sheet_name}'")
                else:
                    print(f"No matching JSON file found for sheet '{sheet_name}' (looking for: {json_filename})")
            
            # Save the updated workbook if any updates were made
            if workbook_updated:
                output_file_path = updated_folder / excel_file.name
                workbook.save(output_file_path)
                print(f"Saved updated workbook: {output_file_path}")
                print(f"Total updates made: {total_updates}")
            else:
                print(f"No JSON files found for workbook '{excel_file.name}' - no updates made")
                
        except Exception as e:
            print(f"Error processing workbook '{excel_file.name}': {e}")
            continue
    
    print(f"\nProcessing complete. Updated workbooks saved to: {updated_folder}")

# Example usage:
# update_excel_from_json('Input_Folder', 'Output_folder')