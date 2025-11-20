"""
Unified Excel Processing System
==============================

This system processes Excel files through a complete pipeline:
1. Excel files (Input_Folder) -> HTML files (html_outputs)
2. HTML files (html_outputs) + Text files (DATA_SOURCES) -> Analysis outputs (Output_folder)

The system is modular, readable, and designed for future extendability.
"""

import openpyxl
from openpyxl.styles import Border, Side
import html
from collections import defaultdict
import argparse
import os
import glob
import csv
import pandas as pd
import json
import re
import urllib.parse
from pathlib import Path
from openai import OpenAI
from dotenv import load_dotenv
from final_json_from_outputFolder_to_xlsx_filling import update_excel_from_json

# Load environment variables
load_dotenv()


class ExcelToHTMLConverter:
    """Converts Excel files to HTML format with proper styling and structure."""
    
    def __init__(self):
        self.border_styles = {
            'thin': '1px SOLID',
            'medium': '2px SOLID', 
            'thick': '3px SOLID',
            'double': '3px DOUBLE',
            'hair': '1px SOLID',
            'dotted': '1px DOTTED',
            'dashed': '1px DASHED',
            'dashDot': '1px DASHED',
            'dashDotDot': '1px DASHED',
            'slantDashDot': '1px DASHED'
        }
    
    def get_border_css(self, border_side):
        """Convert openpyxl border to CSS border string"""
        if not border_side or not border_side.style:
            return 'transparent'
        
        # style = self.border_styles.get(border_side.style)
        style = self.border_styles.get(border_side.style, '1px SOLID')
        color = '#000000'  # Default to black since we're not handling colors
        return f"{style} {color}"
    
    def get_alignment_css(self, alignment):
        """Convert openpyxl alignment to CSS properties"""
        css_props = {}
        
        if alignment:
            # Horizontal alignment
            if alignment.horizontal == 'center':
                css_props['text-align'] = 'center'
            elif alignment.horizontal == 'right':
                css_props['text-align'] = 'right'
            elif alignment.horizontal == 'left':
                css_props['text-align'] = 'left'
            else:
                css_props['text-align'] = 'left'  # default
            
            # Vertical alignment
            if alignment.vertical == 'center':
                css_props['vertical-align'] = 'middle'
            elif alignment.vertical == 'top':
                css_props['vertical-align'] = 'top'
            elif alignment.vertical == 'bottom':
                css_props['vertical-align'] = 'bottom'
            else:
                css_props['vertical-align'] = 'bottom'  # default
            
            # Text wrap
            if alignment.wrap_text:
                css_props['white-space'] = 'normal'
                css_props['overflow'] = 'hidden'
                css_props['word-wrap'] = 'break-word'
            else:
                css_props['white-space'] = 'nowrap'
        
        return css_props
    
    def get_font_css(self, font):
        """Convert openpyxl font to CSS properties"""
        css_props = {}
        
        if font:
            if font.name:
                css_props['font-family'] = font.name
            
            if font.size:
                css_props['font-size'] = f"{font.size}pt"
            
            if font.bold:
                css_props['font-weight'] = 'bold'
        
        return css_props
    
    def generate_cell_style(self, cell):
        """Generate CSS style string for a cell"""

        # Define the default Calibri cell properties
        is_calibri = (
            (not cell.border or (
                (not cell.border.top or cell.border.top.style is None) and
                (not cell.border.bottom or cell.border.bottom.style is None) and
                (not cell.border.left or cell.border.left.style is None) and
                (not cell.border.right or cell.border.right.style is None)
            )) and
            (not cell.alignment or (
                (cell.alignment.horizontal in (None, 'left')) and
                (cell.alignment.vertical in (None, 'bottom')) and
                (not cell.alignment.wrap_text)
            )) and
            (not cell.font or (
                (cell.font.name in (None, 'Calibri')) and
                (cell.font.size in (None, 9, 9.0)) and
                (not cell.font.bold)
            ))
        )
        if is_calibri:
            return 'calibri-cell'
        
        is_calibri_11 = (
            (not cell.border or (
                (not cell.border.top or cell.border.top.style is None) and
                (not cell.border.bottom or cell.border.bottom.style is None) and
                (not cell.border.left or cell.border.left.style is None) and
                (not cell.border.right or cell.border.right.style is None)
            )) and
            (not cell.alignment or (
                (cell.alignment.horizontal in (None, 'left')) and
                (cell.alignment.vertical in (None, 'bottom')) and
                (not cell.alignment.wrap_text)
            )) and
            (not cell.font or (
                (cell.font.name in (None, 'Calibri')) and
                (cell.font.size in (None, 11, 11.0)) and
                (not cell.font.bold)
            ))
        )
        if is_calibri_11:
            return 'calibri-cell-11'

         # Check for bordered Calibri cell (all borders solid 1px black, centered, Calibri 9pt, not bold)
        is_calibri_bordered = (
            cell.border and
            all([
                getattr(cell.border.top, 'style', None) == 'thin',
                getattr(cell.border.bottom, 'style', None) == 'thin',
                getattr(cell.border.left, 'style', None) == 'thin',
                getattr(cell.border.right, 'style', None) == 'thin',
            ]) and
            (not cell.font or (
                (cell.font.name in (None, 'Calibri')) and
                (cell.font.size in (None, 9, 9.0)) and
                (not cell.font.bold)
            )) and
            (cell.alignment and
                cell.alignment.horizontal == 'center' and
                cell.alignment.vertical in (None, 'bottom') and
                not cell.alignment.wrap_text
            )
        )
        if is_calibri_bordered:
            return 'calibri-cell-bordered'
    
        style_props = {}
        
        # Get border styles
        if cell.border:
            if cell.border.top:
                style_props['border-top'] = self.get_border_css(cell.border.top)
            if cell.border.bottom:
                style_props['border-bottom'] = self.get_border_css(cell.border.bottom)
            if cell.border.left:
                style_props['border-left'] = self.get_border_css(cell.border.left)
            if cell.border.right:
                style_props['border-right'] = self.get_border_css(cell.border.right)
        
        # Get alignment styles
        alignment_props = self.get_alignment_css(cell.alignment)
        style_props.update(alignment_props)
        
        # Get font styles
        font_props = self.get_font_css(cell.font)
        style_props.update(font_props)
        
        # Default properties
        style_props.setdefault('color', '#000000')
        style_props.setdefault('direction', 'ltr')
        style_props.setdefault('padding', '0px 3px 0px 3px')
        
        # Convert to CSS string
        return '; '.join([f"{prop}: {value}" for prop, value in style_props.items()])
    
    def find_merged_ranges(self, worksheet):
        """Find all merged cell ranges in the worksheet"""
        merged_ranges = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            
            # Calculate colspan and rowspan
            colspan = max_col - min_col + 1
            rowspan = max_row - min_row + 1
            
            # Mark the top-left cell as the one that should have colspan/rowspan
            merged_ranges[(min_row, min_col)] = {
                'colspan': colspan if colspan > 1 else None,
                'rowspan': rowspan if rowspan > 1 else None,
                'is_merged_root': True
            }
            
            # Mark other cells in the range as hidden
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row != min_row or col != min_col:
                        merged_ranges[(row, col)] = {'hidden': True}
        
        return merged_ranges
    
    def get_column_widths(self, worksheet):
        """Extract column widths from worksheet"""
        column_widths = {}
        for col_letter, dimension in worksheet.column_dimensions.items():
            if dimension.width:
                # Convert Excel width to approximate pixels (rough conversion)
                width_px = int(dimension.width * 7)
                column_widths[col_letter] = f"{width_px}px"
        return column_widths
    
    def get_row_heights(self, worksheet):
        """Extract row heights from worksheet"""
        row_heights = {}
        for row_num, dimension in worksheet.row_dimensions.items():
            if dimension.height:
                height_px = int(dimension.height * 1.33)  # Rough conversion
                row_heights[row_num] = f"{height_px}px"
        return row_heights
    
    def convert_worksheet_to_html(self, worksheet, worksheet_name="Sheet"):
        """Convert a single worksheet to HTML"""
        merged_ranges = self.find_merged_ranges(worksheet)
        column_widths = self.get_column_widths(worksheet)
        row_heights = self.get_row_heights(worksheet)
        
        # Find the actual data range
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 1 and max_col == 1 and not worksheet.cell(1, 1).value:
            return "<p>Empty worksheet</p>"
        
        html_parts = []
        html_parts.append('<div class="ritz grid-container" dir="ltr">')
        html_parts.append('<table class="waffle" cellspacing="0" cellpadding="0">')
        
        # Generate column headers
        html_parts.append('<thead><tr>')
        html_parts.append('<th class="row-header freezebar-origin-ltr"></th>')
        
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            width_style = f' style="width:{column_widths.get(col_letter, "100px")};"' if col_letter in column_widths else ''
            html_parts.append(f'<th id="col{col}"{width_style} class="column-headers-background">{col_letter}</th>')
        
        html_parts.append('</tr></thead>')
        
        # Generate table body
        html_parts.append('<tbody>')
        
        for row in range(1, max_row + 1):
            height_style = f' style="height:{row_heights.get(row, "18px")};"' if row in row_heights else ' style="height:18px"'
            html_parts.append(f'<tr{height_style}>')
            
            # Row header
            html_parts.append(f'<th id="row{row}"{height_style} class="row-headers-background">')
            html_parts.append(f'<div class="row-header-wrapper" style="line-height: 18px">{row}</div>')
            html_parts.append('</th>')
            
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row, col)
                merge_info = merged_ranges.get((row, col), {})
                
                # Skip hidden cells (part of merged ranges)
                if merge_info.get('hidden'):
                    continue
                
                # Get cell value
                cell_value = cell.value
                if cell_value is None:
                    display_value = ''
                elif isinstance(cell_value, (int, float)):
                    display_value = str(cell_value)
                else:
                    display_value = html.escape(str(cell_value))
                
                # Generate cell style
                cell_style = self.generate_cell_style(cell)
                
                # Build cell attributes
                if cell_style == 'calibri-cell':
                    cell_attrs = ['class="calibri-cell"']
                elif cell_style == 'calibri-cell-11':
                    cell_attrs = ['class="calibri-cell-11"']
                elif cell_style == 'calibri-cell-bordered':
                    cell_attrs = ['class="calibri-cell-bordered"']
                else:
                    cell_attrs = [f'style="{cell_style}"']
                
                if merge_info.get('colspan'):
                    cell_attrs.append(f'colspan="{merge_info["colspan"]}"')
                if merge_info.get('rowspan'):
                    cell_attrs.append(f'rowspan="{merge_info["rowspan"]}"')
                
                attrs_str = ' ' + ' '.join(cell_attrs) if cell_attrs else ''
                
                html_parts.append(f'<td{attrs_str}>{display_value}</td>')
            
            html_parts.append('</tr>')
        
        html_parts.append('</tbody>')
        html_parts.append('</table>')
        html_parts.append('</div>')
        
        return '\n'.join(html_parts)
    
    def generate_css(self):
        """Generate basic CSS for the HTML table"""
        css = """
        <style type="text/css">
            .ritz .waffle a {
                color: inherit;
            }
            .ritz .waffle {
                border-collapse: collapse;
                font-family: Arial, sans-serif;
                font-size: 10pt;
            }
            .ritz .waffle td, .ritz .waffle th {
                padding: 3px;
                border: 1px solid #ccc;
            }
            .column-headers-background {
                background-color: #f0f0f0;
                font-weight: bold;
            }
            .row-headers-background {
                background-color: #f0f0f0;
                font-weight: bold;
                text-align: center;
            }
            .row-header-wrapper {
                line-height: 18px;
            }

            .calibri-cell {
                border-top: none;
                border-bottom: none;
                border-left: none;
                border-right: none;
                text-align: left;
                vertical-align: bottom;
                white-space: nowrap;
                font-family: Calibri, Arial, sans-serif;
                font-size: 9pt;
                color: #000000;
                direction: ltr;
                padding: 0px 3px 0px 3px;
            }

            .calibri-cell-bordered {
                border-top: 1px solid #000000;
                border-bottom: 1px solid #000000;
                border-left: 1px solid #000000;
                border-right: 1px solid #000000;
                text-align: center;
                vertical-align: bottom;
                white-space: nowrap;
                font-family: Calibri, Arial, sans-serif;
                font-size: 9pt;
                color: #000000;
                direction: ltr;
                padding: 0px 3px 0px 3px;
            }

            .calibri-cell-11 {
                border-top: none;
                border-bottom: none;
                border-left: none;
                border-right: none;
                text-align: left;
                vertical-align: bottom;
                white-space: nowrap;
                font-family: Calibri, sans-serif;
                font-size: 11pt;
                color: #000000;
                direction: ltr;
                padding: 0px 3px 0px 3px;
            }
        </style>
        """
        return css
    
    def convert_worksheet_to_separate_html(self, worksheet, worksheet_name, base_filename, output_folder):
        """Convert a single worksheet to a separate HTML file"""
        html_parts = []
        html_parts.append('<!DOCTYPE html>')
        html_parts.append('<html>')
        html_parts.append('<head>')
        html_parts.append('<meta http-equiv="Content-Type" content="text/html; charset=utf-8">')
        html_parts.append(f'<title>{html.escape(worksheet_name)} - {html.escape(base_filename)}</title>')
        html_parts.append(self.generate_css())
        html_parts.append('</head>')
        html_parts.append('<body>')
        
        # Add worksheet title
        html_parts.append(f'<h1>Sheet: {html.escape(worksheet_name)}</h1>')
        html_parts.append(self.convert_worksheet_to_html(worksheet, worksheet_name))
        
        html_parts.append('</body>')
        html_parts.append('</html>')
        
        # Create filename: base_filename_worksheet_name.html
        safe_worksheet_name = "".join(c for c in worksheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_worksheet_name = safe_worksheet_name.replace(' ', '_')
        output_filename = f"{base_filename}_{safe_worksheet_name}.html"
        output_file_path = os.path.join(output_folder, output_filename)
        
        # Write to file
        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html_parts))
        
        print(f"✓ Converted worksheet '{worksheet_name}' to {output_filename}")
        return output_file_path
    
    def process_all_excel_files(self, input_folder, output_folder):
        """Process all Excel files in input folder and save HTML files to output folder"""
        # Ensure output folder exists
        os.makedirs(output_folder, exist_ok=True)
        
        # Find all Excel files in input folder
        excel_pattern = os.path.join(input_folder, "*.xlsx")
        excel_files = glob.glob(excel_pattern)
        
        if not excel_files:
            print(f"No Excel files found in {input_folder}")
            return []
        
        print(f"Found {len(excel_files)} Excel files to process:")
        for file in excel_files:
            print(f"  - {os.path.basename(file)}")
        
        processed_files = []
        
        for excel_file in excel_files:
            try:
                # Load workbook
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
                
                # Get base filename (without extension)
                base_name = os.path.splitext(os.path.basename(excel_file))[0]
                
                print(f"\nProcessing {base_name} with {len(workbook.sheetnames)} worksheet(s):")
                
                # Process each worksheet separately
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    output_file = self.convert_worksheet_to_separate_html(worksheet, sheet_name, base_name, output_folder)
                    processed_files.append(output_file)
                
            except Exception as e:
                print(f"Error processing {excel_file}: {e}")
                continue
        
        print(f"\n✓ Excel to HTML conversion complete! {len(processed_files)} HTML files created.")
        return processed_files


class HTMLAnalyzer:
    """Analyzes HTML files with corresponding text files using OpenAI."""
    
    def __init__(self):
        try:
            self.client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        except Exception as e:
            print(f"Warning: Could not initialize OpenAI client: {e}")
            self.client = None
    
    def save_as_csv(self, cell_mappings, csv_path, base_name):
        """Save cell mappings as CSV file for easy import."""
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['row', 'column', 'cell_reference', 'value', 'context', 'source_file']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for mapping in cell_mappings:
                row_data = mapping.copy()
                row_data['source_file'] = base_name
                writer.writerow(row_data)
    
    def save_as_excel(self, cell_mappings, excel_path, base_name):
        """Save cell mappings as Excel file."""
        try:
            # Create DataFrame
            df = pd.DataFrame(cell_mappings)
            df['source_file'] = base_name
            
            # Reorder columns
            column_order = ['row', 'column', 'cell_reference', 'value', 'context', 'source_file']
            df = df[column_order]
            
            # Save to Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Cell_Mappings', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Cell_Mappings']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
        except Exception as e:
            print(f"Warning: Could not create Excel file {excel_path}: {e}")
            # Fallback: create a simple CSV
            csv_path = excel_path.replace('.xlsx', '.csv')
            self.save_as_csv(cell_mappings, csv_path, base_name)
    
    def count_html_rows(self, html_path):
        """
        Count the number of data rows in an HTML file by parsing row IDs.
        
        Args:
            html_path (str): Path to HTML file
            
        Returns:
            int: Number of data rows (excluding header)
        """
        # Convert file:// URL to local path if needed
        if html_path.startswith("file:///"):
            html_path = urllib.parse.unquote(html_path[8:])
        
        # Read HTML file with multiple encoding attempts
        html_content = None
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
        
        for encoding in encodings_to_try:
            try:
                with open(html_path, 'r', encoding=encoding) as file:
                    html_content = file.read()
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if html_content is None:
            raise ValueError(f"Could not read {html_path} with any of the attempted encodings: {encodings_to_try}")
        
        # Find all row IDs using regex
        row_pattern = r'id="row(\d+)"'
        row_matches = re.findall(row_pattern, html_content)
        
        if not row_matches:
            return 0
        
        # Convert to integers and find the maximum
        row_numbers = [int(row_id) for row_id in row_matches]
        max_row = max(row_numbers)
        
        return max_row

    def process_html_file_in_chunks(self, html_path, text_file_path, output_path, chunk_size=30):
        """
        Process HTML file in chunks of specified size, updating JSON incrementally.
        
        Args:
            html_path (str): Path to HTML file
            text_file_path (str): Path to corresponding text file
            output_path (str): Path to save the analysis results
            chunk_size (int): Number of rows to process in each chunk
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Count total rows in HTML file
            total_rows = self.count_html_rows(html_path)
            print(f"  📊 Total rows in HTML file: {total_rows}")
            
            if total_rows == 0:
                print("  ⚠️  No rows found in HTML file")
                return False
            
            # Initialize empty cell mappings list
            all_cell_mappings = []
            base_name = os.path.splitext(os.path.basename(html_path))[0]
            
            # Calculate number of chunks needed
            num_chunks = (total_rows + chunk_size - 1) // chunk_size  # Ceiling division
            print(f"  🔄 Processing in {num_chunks} chunks of {chunk_size} rows each")
            
            # Process each chunk
            for chunk_num in range(num_chunks):
                row_start = (chunk_num * chunk_size) + 1
                row_end = min((chunk_num + 1) * chunk_size, total_rows)
                
                print(f"  📝 Processing chunk {chunk_num + 1}/{num_chunks}: rows {row_start}-{row_end}")
                
                # Create temporary output path for this chunk
                chunk_output_path = output_path.replace('.txt', f'_chunk_{chunk_num + 1}.txt')
                
                # Process this chunk
                success = self.analyze_html_table_with_openai(
                    html_path, text_file_path, chunk_output_path, 
                    row_start=row_start, row_end=row_end
                )
                
                if success:
                    # Load the chunk results
                    chunk_json_path = chunk_output_path.replace('.txt', '.json')
                    if os.path.exists(chunk_json_path):
                        with open(chunk_json_path, 'r', encoding='utf-8') as f:
                            chunk_mappings = json.load(f)
                        all_cell_mappings.extend(chunk_mappings)
                        print(f"    ✓ Chunk {chunk_num + 1} completed: {len(chunk_mappings)} cells found")
                        
                        # Clean up temporary chunk files
                        try:
                            os.remove(chunk_output_path)
                            os.remove(chunk_json_path)
                        except:
                            pass
                    else:
                        print(f"    ⚠️  No JSON file created for chunk {chunk_num + 1}")
                else:
                    print(f"    ✗ Chunk {chunk_num + 1} failed")
            
            # Save final combined results
            if all_cell_mappings:
                print(f"  💾 Saving combined results: {len(all_cell_mappings)} total cells")
                
                # Create structured output files
                csv_path = output_path.replace('.txt', '.csv')
                self.save_as_csv(all_cell_mappings, csv_path, base_name)
                
                excel_path = output_path.replace('.txt', '.xlsx')
                self.save_as_excel(all_cell_mappings, excel_path, base_name)
                
                # Save as JSON for reference
                json_path = output_path.replace('.txt', '.json')
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(all_cell_mappings, f, indent=2, ensure_ascii=False)
                
                # Save summary text file
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(f"HTML File: {os.path.basename(html_path)}\n")
                    f.write(f"Text File: {os.path.basename(text_file_path)}\n")
                    f.write(f"Total Rows Processed: {total_rows}\n")
                    f.write(f"Chunks Processed: {num_chunks}\n")
                    f.write("=" * 80 + "\n\n")
                    f.write(f"Found {len(all_cell_mappings)} cells to fill:\n\n")
                    for mapping in all_cell_mappings:
                        f.write(f"Cell {mapping['cell_reference']}: {mapping['value']}\n")
                        f.write(f"  Context: {mapping['context']}\n\n")
                    f.write(f"\nFiles generated:\n")
                    f.write(f"- {os.path.basename(csv_path)} (CSV for import)\n")
                    f.write(f"- {os.path.basename(excel_path)} (Excel file)\n")
                    f.write(f"- {os.path.basename(json_path)} (JSON data)\n")
                
                print(f"  ✓ Successfully processed {os.path.basename(html_path)} in {num_chunks} chunks")
                return True
            else:
                print(f"  ⚠️  No cells found for {os.path.basename(html_path)}")
                return False
                
        except Exception as e:
            print(f"  ✗ Error processing {os.path.basename(html_path)} in chunks: {e}")
            return False

    def analyze_html_table_with_openai(self, html_path, text_file_path, output_path, row_start=1, row_end=None):
        """
        Analyze HTML table and match data from text file, save results to output file.
        
        Args:
            html_path (str): Path to HTML file
            text_file_path (str): Path to corresponding text file
            output_path (str): Path to save the analysis results
            row_start (int): Starting row number (1-based)
            row_end (int): Ending row number (1-based, inclusive). If None, processes all rows.
        """
        # Convert file:// URL to local path if needed
        if html_path.startswith("file:///"):
            html_path = urllib.parse.unquote(html_path[8:])  # Remove file:/// and decode
        
        # Read HTML file with multiple encoding attempts
        html_content = None
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
        
        for encoding in encodings_to_try:
            try:
                with open(html_path, 'r', encoding=encoding) as file:
                    html_content = file.read()
                print(f"  ✓ Read {os.path.basename(html_path)} with {encoding} encoding")
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if html_content is None:
            raise ValueError(f"Could not read {html_path} with any of the attempted encodings: {encodings_to_try}")
        
        # Read text file with encoding handling
        text_data = None
        for encoding in encodings_to_try:
            try:
                with open(text_file_path, 'r', encoding=encoding) as file:
                    text_data = file.read().strip()
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if text_data is None:
            raise ValueError(f"Could not read {text_file_path} with any of the attempted encodings: {encodings_to_try}")
        
        # Create detailed prompt for table analysis with structured output
        row_range_text = f"Process rows {row_start} to {row_end}" if row_end else f"Process from row {row_start} to the end"
        
        prompt = f"""
        Analyze the provided HTML table structure and available text data to intelligently determine which cells require updating or filling.

## INPUT:
- **HTML Table Content:** {html_content}
- **Text Data for Filling:** {text_data}
- **Row Range:** {row_range_text}

## TASK:
You are an advanced table analysis agent. Your goal is to:

1.  **Parse the HTML Table:**
    *   Identify the complete structure: rows, columns, headers (both column headers and row headers/labels).
    *   Map this structure to a spreadsheet-like coordinate system (A1, B2, etc.), where columns are letters (A, B, ..., Z, AA, AB, ...) and rows are numbers starting from 1.
    *   **IMPORTANT:** Analyze the complete table structure for full context, but only return all the cells that need to be filled from rows {row_start} to {row_end if row_end else 'the end of the table'} in your output.

2.  **Identify Target Cells for Potential Update:**
    *   Your objective is to find cells that **should be updated or filled** with the provided `text_data`. This includes, but is not limited to:
        *   **a) Obviously Empty Cells:** Truly empty cells (`<td></td>`), cells with only whitespace, non-breaking spaces (`&nbsp;`), or generic placeholders ("___", "???", "N/A", "TBD", "Pending", "Not specified").
        *   **b) Instructional Placeholders:** Cells containing text that explicitly or implicitly instructs a user to fill them (e.g., "Enter value here", "Fill in", "Provide data", "Type name", "here it should be merchant", "insert total").
        *   **c) Incorrect or Incomplete Data:** Cells containing data that is clearly a temporary substitute, incorrect, or incomplete based on the context of its row and column. This includes:
            *   Generic stand-ins like "-", "?", "*", "xxx".
            *   Data that semantically conflicts with the column header or row label.
            *   Data that is logically inconsistent with surrounding values (e.g., a blank in a sequence of numbers, a word in a numeric column).
        *   **d) Default or Example Values:** Cells containing values that appear to be examples or defaults that need to be replaced with real data (e.g., "Sample Text", "Jane Doe", "100.00").

3.  **Perform Deep Contextual Analysis & Data Matching:**
    *   For each cell identified for potential update, perform a deep analysis:
        *   **Column Context:** The header of the column it belongs to (data category, expected data type: text, number, date, percentage, currency).
        *   **Row Context:** The label or content of the row it belongs to, as well as adjacent cell values. Look for patterns.
        *   **Semantic Context:** The overall meaning and purpose of the table. Does the existing cell content fulfill that purpose?
        *   **Data Comparison:** Does the provided `text_data` contain a value that is a *better fit* for this cell's context than its current content?
    *   Match the provided `text_data` to these cells by finding the most appropriate fit. Prioritize semantic alignment and data type compatibility. The goal is to *improve* the table's accuracy and completeness.

4.  **Generate Structured Output:**
    *   Produce a JSON array containing an object for each cell that should be updated.
    *   Each object MUST have the following structure:
            [
                {{
                    "cell_reference": "B3", // The Excel-style cell coordinate
                    "value": "Correct Data", // The specific string from `text_data` to insert
                    "context": "Brief description of why this cell should be updated. Reference column/row headers and explain why the current content is insufficient/incorrect and why the new value is correct.", // Explain your reasoning
                }},
                {{
                    "cell_reference": "B2", 
                    "value": "Another data value",
                    "context": "Brief description"
                }}
            ]
    *   **Only include cells** where a superior match from the `text_data` exists and the update is justified.
    *   **Do not include** cells that are already correct and complete.
    *   **Analyze the complete table structure for full context, but only return all the cells that need to be filled from rows {row_start} to {row_end if row_end else 'the end'} in your output.**

## OUTPUT REQUIREMENTS:
*   Output MUST be **ONLY** the JSON array, with no additional text, commentary, or formatting before or after.
*   The JSON must be perfectly formatted and valid.
*   Use double quotes for all JSON strings.

## AGENT INSTRUCTIONS:
*   **Think Critically:** You must not just find blanks. You must evaluate the *quality* and *appropriateness* of existing cell content. Ask yourself: "Is this the final, correct data for this cell?"
*   **Justify Updates:** The burden of proof is on you. Only recommend an update if you can clearly articulate why the new value is better than the old one. The `context` field is crucial.
*   **Leverage Patterns:** Use the structure of the table (headers, data types, consistent formatting in columns) as a powerful guide for what belongs in a cell.
*   **Full Context Analysis:** Analyze the complete table structure for full context, but only return all the cells that need to be filled from rows {row_start} to {row_end if row_end else 'end'} in your output.
*   **DO NOT SKIP ANY ROWS OR CELLS WITHIN THIS RANGE**
"""
        
        try:
            # Check if OpenAI client is available
            if self.client is None:
                raise Exception("OpenAI client not initialized. Please check your API key and dependencies.")
            
            # Make API call to OpenAI
            response = self.client.chat.completions.create(
                model="gpt-5-mini",
                messages=[
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                max_completion_tokens=20000  # Increased for comprehensive analysis
            )
            
            # Get the analysis content
            analysis_content = response.choices[0].message.content
            
            # Try to parse JSON response
            try:
                # Extract JSON from response (in case there's extra text)
                json_match = re.search(r'\[.*\]', analysis_content, re.DOTALL)
                if json_match:
                    json_str = json_match.group(0)
                    cell_mappings = json.loads(json_str)
                else:
                    # Try to parse the entire response as JSON
                    cell_mappings = json.loads(analysis_content)
                
                # Create structured output files
                base_name = os.path.splitext(os.path.basename(html_path))[0]
                
                # Save as CSV
                csv_path = output_path.replace('.txt', '.csv')
                self.save_as_csv(cell_mappings, csv_path, base_name)
                
                # Save as Excel
                excel_path = output_path.replace('.txt', '.xlsx')
                self.save_as_excel(cell_mappings, excel_path, base_name)
                
                # Save as JSON for reference
                json_path = output_path.replace('.txt', '.json')
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(cell_mappings, f, indent=2, ensure_ascii=False)

                

                
                # Save summary text file
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(f"HTML File: {os.path.basename(html_path)}\n")
                    f.write(f"Text File: {os.path.basename(text_file_path)}\n")
                    f.write("=" * 80 + "\n\n")
                    f.write(f"Found {len(cell_mappings)} cells to fill:\n\n")
                    for mapping in cell_mappings:
                        f.write(f"Cell {mapping['cell_reference']}: {mapping['value']}\n")
                        f.write(f"  Context: {mapping['context']}\n\n")
                    f.write(f"\nFiles generated:\n")
                    f.write(f"- {os.path.basename(csv_path)} (CSV for import)\n")
                    f.write(f"- {os.path.basename(excel_path)} (Excel file)\n")
                    f.write(f"- {os.path.basename(json_path)} (JSON data)\n")
                
                print(f"  ✓ Generated analysis files for {os.path.basename(html_path)}")
                return True
                
            except (json.JSONDecodeError, KeyError) as e:
                print(f"  ✗ Error parsing JSON response for {os.path.basename(html_path)}: {e}")
                # Fallback: save raw response
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(f"HTML File: {os.path.basename(html_path)}\n")
                    f.write(f"Text File: {os.path.basename(text_file_path)}\n")
                    f.write("=" * 80 + "\n\n")
                    f.write("Raw AI Response (JSON parsing failed):\n\n")
                    f.write(analysis_content)
                return False
            
        except Exception as e:
            print(f"  ✗ Error analyzing {html_path}: {e}")
            return False
    
    def find_matching_files(self, html_outputs_dir, data_sources_dir):
        """
        Find matching HTML and TXT file pairs.
        
        Args:
            html_outputs_dir (str): Directory containing HTML files
            data_sources_dir (str): Directory containing TXT files
        
        Returns:
            list: List of tuples (html_path, txt_path, output_path)
        """
        matches = []
        
        # Get all HTML files from html_outputs
        html_files = glob.glob(os.path.join(html_outputs_dir, "*.html"))
        
        for html_path in html_files:
            html_filename = os.path.basename(html_path)
            # Remove .html extension to get base name
            base_name = os.path.splitext(html_filename)[0]
            
            # Look for corresponding TXT file in DATA_SOURCES
            txt_path = os.path.join(data_sources_dir, f"{base_name}.txt")
            
            if os.path.exists(txt_path):
                # Create output path in Output_folder
                output_path = os.path.join("Output_folder", f"{base_name}.txt")
                matches.append((html_path, txt_path, output_path))
            else:
                print(f"  ⚠ Warning: No matching TXT file found for {html_filename}")
        
        return matches
    
    def process_all_files(self):
        """
        Process all HTML files from html_outputs folder with their corresponding TXT files.
        """
        # Define directories
        html_outputs_dir = "html_outputs"
        data_sources_dir = "DATA_SOURCES"
        output_dir = "Output_folder"
        
        # Check if directories exist
        if not os.path.exists(html_outputs_dir):
            print(f"Error: {html_outputs_dir} directory not found")
            return False
        
        if not os.path.exists(data_sources_dir):
            print(f"Error: {data_sources_dir} directory not found")
            return False
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Check if OpenAI API key is set
        if not os.getenv('OPENAI_API_KEY'):
            print("Error: Please set your OPENAI_API_KEY environment variable")
            return False
        
        # Find matching files
        matches = self.find_matching_files(html_outputs_dir, data_sources_dir)
        
        if not matches:
            print("No matching HTML and TXT file pairs found")
            return False
        
        print(f"Found {len(matches)} matching file pairs to process:")
        for html_path, txt_path, output_path in matches:
            print(f"  {os.path.basename(html_path)} + {os.path.basename(txt_path)} -> {os.path.basename(output_path)}")
        
        print("\nStarting HTML analysis...")
        
        # Process each file pair
        successful = 0
        failed = 0
        
        for i, (html_path, txt_path, output_path) in enumerate(matches, 1):
            print(f"\n[{i}/{len(matches)}] Processing: {os.path.basename(html_path)}")
            
            try:
                # Use chunked processing instead of single file processing
                success = self.process_html_file_in_chunks(html_path, txt_path, output_path, chunk_size=30)
                if success:
                    successful += 1
                    print(f"  ✓ Successfully processed: {os.path.basename(html_path)}")
                else:
                    failed += 1
                    print(f"  ✗ Failed to process: {os.path.basename(html_path)}")
            except Exception as e:
                failed += 1
                print(f"  ✗ Error processing {os.path.basename(html_path)}: {e}")
        
        print(f"\n✓ HTML Analysis complete!")
        print(f"  Successfully processed: {successful} files")
        print(f"  Failed: {failed} files")
        print(f"  Results saved in: {output_dir}")
        
        return successful > 0


class UnifiedProcessor:
    """Main processor that orchestrates the entire workflow."""
    
    def __init__(self):
        self.excel_converter = ExcelToHTMLConverter()
        self.html_analyzer = HTMLAnalyzer()
    
    def process_complete_workflow(self, input_folder="Input_Folder", 
                                 html_outputs_folder="html_outputs", 
                                 data_sources_folder="DATA_SOURCES", 
                                 output_folder="Output_folder"):
        """
        Process the complete workflow:
        1. Excel files -> HTML files
        2. HTML files + Text files -> Analysis outputs
        """
        print("=" * 60)
        print("UNIFIED EXCEL PROCESSING SYSTEM")
        print("=" * 60)
        print("Processing workflow:")
        print("1. Excel files (Input_Folder) -> HTML files (html_outputs)")
        print("2. HTML files (html_outputs) + Text files (DATA_SOURCES) -> Analysis outputs (Output_folder)")
        print("=" * 60)
        
        # Step 1: Convert Excel files to HTML
        print("\n🔄 STEP 1: Converting Excel files to HTML...")
        print("-" * 40)
        
        if not os.path.exists(input_folder):
            print(f"Error: Input folder '{input_folder}' not found")
            return False
        
        html_files = self.excel_converter.process_all_excel_files(input_folder, html_outputs_folder)
        
        if not html_files:
            print("No Excel files were processed. Exiting.")
            return False
        
        # Step 2: Analyze HTML files with text data
        print("\n🔄 STEP 2: Analyzing HTML files with text data...")
        print("-" * 40)
        
        success = self.html_analyzer.process_all_files()
        
        if success:
            print("\n" + "=" * 60)
            print("✅ COMPLETE WORKFLOW SUCCESSFUL!")
            print("=" * 60)
            print(f"✓ Converted {len(html_files)} Excel files to HTML")
            print(f"✓ Generated analysis files in {output_folder}")
            print("=" * 60)
        else:
            print("\n" + "=" * 60)
            print("❌ WORKFLOW COMPLETED WITH ISSUES")
            print("=" * 60)
            print("Check the output above for any errors.")
            print("=" * 60)
        
        return success


def main():
    """Main function to run the unified processor."""
    processor = UnifiedProcessor()
    
    # Process the complete workflow
    success = processor.process_complete_workflow()

    if success:
        print("\n🎉 All AI processing completed successfully!")
    else:
        print("\n⚠️  Processing completed with some issues. Please check the output above.")

    input_folder_name = os.path.join(os.path.dirname(__file__), "Input_Folder")
    output_folder_name = os.path.join(os.path.dirname(__file__), "Output_folder")
    update_excel_from_json(input_folder_name, output_folder_name)
    
    print("\n🎉 All Excel updating completed successfully!")


if __name__ == "__main__":
    # Required installation:
    # pip install openai python-dotenv openpyxl pandas
    
    # Set your OpenAI API key:
    # export OPENAI_API_KEY="your-api-key-here"
    # Or create a .env file with: OPENAI_API_KEY=your-api-key-here
    
    main()
